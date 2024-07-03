import openpyxl

def read_excel_columns(file_path, column1, column2, column3, min_row, max_row):
    # Откроем файл Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Получим значения из указанных столбцов и строк
    column1_values = []
    column2_values = []
    column3_values = []

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min(column1, column2, column3), max_col=max(column1, column2, column3), values_only=True):
        if row:
            if column1 - 1 < len(row):
                column1_values.append(row[column1-1])
            else:
                column1_values.append(None)

            if column2 - 1 < len(row):
                column2_values.append(row[column2-1])
            else:
                column2_values.append(None)
            
            if column3 - 1 < len(row):
                column3_values.append(row[column3-1])
            else:
                column3_values.append(None)

    return column1_values, column2_values, column3_values

def count_colored_cells(graph_path):
    # Откроем файл Excel
    workbook = openpyxl.load_workbook(graph_path)
    sheet = workbook.active

    # Получим заголовки (наименования) из первой строки
    headers = [cell.value for cell in sheet[2]]

    green_counts = [0] * len(headers)
    blue_counts = [0] * len(headers)

    # Цвета, которые мы ищем
    green_rgb = '99CC00'
    blue_rgb = '3366FF'

    # Пройдем по всем строкам начиная со второй
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for idx, cell in enumerate(row):
            if cell.fill and cell.fill.start_color and cell.fill.start_color.index:
                cell_color = cell.fill.start_color.index
                if cell_color == green_rgb:
                    green_counts[idx] += 1
                elif cell_color == blue_rgb:
                    blue_counts[idx] += 1

    return headers, green_counts, blue_counts

def compare_groups(column1_values, column2_values, column3_values, headers, green_counts, blue_counts):
    
    # Итерируемся по массивам
    for i in range(len(column1_values)):
        if column1_values[i] == headers[i]:
            print(f"Наименование {column1_values[i]} совпадает:")
            print(f" Параметр 1: Group1[{i}] = {column2_values[i]}, Group2[{i}] = {green_counts[i]}")
            print(f" Параметр 2: Group1[{i}] = {column3_values[i]}, Group2[{i}] = {blue_counts[i]}")
        else:
            print(f"Наименование {i+1} не совпадает: Group1[{i}] = {column1_values[i]}, Group2[{i}] = {column1_values[i]}")

def main():
    #данные основного файла
    file_path = "606.xlsx" 
    column1 = 14  # Номер первого столбца (например, 1 для столбца A)
    column2 = 15  # Номер второго столбца (например, 2 для столбца B)
    column3 = 17  # Номер третьего столбца (например, 3 для столбца C)
    min_row = 27  # Начальная строка
    max_row = 74  # Конечная строка

    # Получим данные из столбцов
    column1_values, column2_values, column3_values = read_excel_columns(file_path, column1, column2, column3, min_row, max_row)

    #данные графика
    graph_path = "606(1).xlsx" 

    # Получим данные из таблицы
    headers, green_counts, blue_counts = count_colored_cells(file_path)

    # Тестовый вывод массивов
    print("Время:", column1_values)
    print("Прямой путь:", column2_values)
    print("Обратный путь:", column3_values)
    print("Наименования:", headers)
    print("Количество зеленых клеток:", green_counts)
    print("Количество синих клеток:", blue_counts)

    compare_groups(column1_values, column2_values, column3_values, headers, green_counts, blue_counts)

if __name__ == "__main__":
    main()